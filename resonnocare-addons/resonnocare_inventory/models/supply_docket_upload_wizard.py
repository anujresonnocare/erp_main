# -*- coding: utf-8 -*-
import base64
from datetime import datetime
from io import BytesIO

from odoo import _, fields, models
from odoo.exceptions import UserError


class ResonnocareSupplyDocketUploadWizard(models.TransientModel):
    _name = "resonnocare.supply.docket.upload.wizard"
    _description = "Bulk Docket Upload for Supply Orders"

    file_data = fields.Binary(string="Excel File", required=True)
    file_name = fields.Char(string="Filename")
    update_count = fields.Integer(string="Updated Orders", readonly=True)

    def _normalize(self, value):
        return str(value or "").strip().lower().replace(" ", "").replace("_", "")

    def _first_present_index(self, indexes, *keys):
        for key in keys:
            if key in indexes:
                return indexes[key]
        return None

    def _parse_date(self, value):
        if not value:
            return False
        if isinstance(value, datetime):
            return value.date()
        if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
            return value
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except Exception:
                continue
        return False

    def action_apply_upload(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError(_("Please upload an Excel file."))
        try:
            import openpyxl
        except Exception:
            raise UserError(_("Excel upload requires python package 'openpyxl' on the server."))

        try:
            wb = openpyxl.load_workbook(BytesIO(base64.b64decode(self.file_data)), data_only=True)
        except Exception as err:
            raise UserError(_("Unable to read Excel file. Error: %s") % err)

        ws = wb.active
        if ws.max_row < 2:
            raise UserError(_("Excel file does not contain data rows."))

        header_row = [self._normalize(c.value) for c in ws[1]]
        idx = {name: i for i, name in enumerate(header_row)}

        key_order = self._first_present_index(idx, "orderforsupplyno", "orderno")
        key_transfer = self._first_present_index(idx, "transferno", "pickingno", "name")
        key_grn = self._first_present_index(idx, "grnnumber", "stnnumber")
        key_courier = self._first_present_index(idx, "couriername")
        key_docket = self._first_present_index(idx, "docketnumber")
        key_docket_date = self._first_present_index(idx, "docketdate(yyyy-mm-dd)", "docketdate")
        key_remarks = self._first_present_index(idx, "dispatchremarks")

        if key_courier is None and key_docket is None and key_docket_date is None and key_remarks is None:
            raise UserError(
                _(
                    "No supported update columns found. Expected at least one of: "
                    "Courier Name, Docket Number, Docket Date, Dispatch Remarks."
                )
            )
        if key_order is None and key_transfer is None and key_grn is None:
            raise UserError(
                _(
                    "No identifier column found. Expected one of: "
                    "Order For Supply No / Transfer No / GRN Number."
                )
            )

        updates = 0
        picking_model = self.env["stock.picking"].sudo()
        for r in range(2, ws.max_row + 1):
            row = ws[r]
            order_no = str(row[key_order].value).strip() if key_order is not None and row[key_order].value else False
            transfer_no = str(row[key_transfer].value).strip() if key_transfer is not None and row[key_transfer].value else False
            grn_no = str(row[key_grn].value).strip() if key_grn is not None and row[key_grn].value else False

            if not any([order_no, transfer_no, grn_no]):
                continue

            domain = [("picking_type_code", "=", "internal"), ("is_clinic_supply", "=", True)]
            if order_no:
                domain.append(("name", "=", order_no))
            elif transfer_no:
                domain.append(("name", "=", transfer_no))
            elif grn_no:
                domain.append(("stn_number", "=", grn_no))

            picking = picking_model.search(domain, limit=1)
            if not picking:
                continue

            vals = {}
            if key_courier is not None and row[key_courier].value:
                vals["courier_name"] = str(row[key_courier].value).strip()
            if key_docket is not None and row[key_docket].value:
                vals["docket_number"] = str(row[key_docket].value).strip()
            if key_docket_date is not None and row[key_docket_date].value:
                parsed_date = self._parse_date(row[key_docket_date].value)
                if parsed_date:
                    vals["docket_date"] = parsed_date
            if key_remarks is not None and row[key_remarks].value:
                vals["dispatch_remarks"] = str(row[key_remarks].value).strip()

            if vals:
                picking.write(vals)
                updates += 1

        self.update_count = updates
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "type": "success",
                "title": _("Bulk Docket Upload"),
                "message": _("%s order(s) updated successfully.") % updates,
                "sticky": False,
            },
        }
